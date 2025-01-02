import tkinter as tk
from tkinter import *
from tkinter import ttk
from tkinter import messagebox
from PIL import Image, ImageTk
import pymysql
import openpyxl
from openpyxl.utils import get_column_letter

mydb = pymysql.connect(host="localhost",
    user="root",
    password="ex.-1420",
    database="nuclear_waste_storage_registry")

cursor = mydb.cursor()


def noc_add(idd, name, owner):
    cursor.execute(f"INSERT INTO nws_owner_company (CompanyID, CompanyName, CompanyOwner) "
    f"VALUES ({idd}, '{name}', '{owner}');")
    mydb.commit()

def noc_update(idd, name=None, owner=None):
    fields_to_update = []
    values = []

    if name:
        fields_to_update.append("CompanyName = %s")
        values.append(name)

    if owner:
        fields_to_update.append("CompanyOwner = %s")
        values.append(owner)

    if not fields_to_update:
        return

    values.append(idd)
    query = f"UPDATE nws_owner_company SET {', '.join(fields_to_update)} WHERE CompanyID = %s;"
    cursor.execute(query, values)
    mydb.commit()

def noc_delete(idd):
    cursor.execute(f"DELETE FROM nws_owner_company WHERE CompanyID={idd};")
    mydb.commit()

def noc_get_all():
    cursor.execute(f"SELECT * FROM nws_owner_company;")
    return cursor.fetchall()

def lsa_add(idd, name, owner):
    cursor.execute(f"INSERT INTO licensed_servicing_agency (AgencyID, AgencyName, AgencyOwner) "
    f"VALUES ({idd}, '{name}', '{owner}');")
    mydb.commit()

def lsa_update(idd, name=None, owner=None):
    fields_to_update = []
    values = []

    if name:
        fields_to_update.append("AgencyName = %s")
        values.append(name)

    if owner:
        fields_to_update.append("AgencyOwner = %s")
        values.append(owner)

    if not fields_to_update:
        return

    values.append(idd)
    query = f"UPDATE licensed_servicing_agency SET {', '.join(fields_to_update)} WHERE AgencyID = %s;"
    cursor.execute(query, values)
    mydb.commit()

def lsa_delete(idd):
    cursor.execute(f"DELETE FROM licensed_servicing_agency WHERE AgencyID={idd};")
    mydb.commit()

def lsa_get_all():
    cursor.execute(f"SELECT * FROM licensed_servicing_agency;")
    return cursor.fetchall()

def nl_add(idd, reg, dist, lat, long):
    cursor.execute(f"INSERT INTO nws_location (LocationID, Region, District, Latitude, Longitude) "
                   f"VALUES ({idd}, '{reg}', '{dist}', {lat}, {long});")
    mydb.commit()

def nl_update(idd, reg=None, dist=None, lat=None, long=None):
    fields_to_update = []
    values = []

    if reg:
        fields_to_update.append("Region = %s")
        values.append(reg)

    if dist:
        fields_to_update.append("District = %s")
        values.append(dist)

    if lat:
        fields_to_update.append("Latitude = %s")
        values.append(lat)

    if long:
        fields_to_update.append("Longitude = %s")
        values.append(long)

    if not fields_to_update:
        return

    values.append(idd)
    query = f"UPDATE nws_location SET {', '.join(fields_to_update)} WHERE LocationID = %s;"
    cursor.execute(query, values)
    mydb.commit()

def nl_delete(idd):
    cursor.execute(f"DELETE FROM nws_location WHERE LocationID={idd};")
    mydb.commit()

def nl_get_all():
    cursor.execute(f"SELECT * FROM nws_location;")
    return cursor.fetchall()

def sc_add(idd, state, exp, ag_id):
    if exp:
        cursor.execute(f"INSERT INTO storage_certificate (CertificateID, OverallStorageState, ExpirationDate, ServicingAgencyID) "
                       f"VALUES ({idd}, '{state}', '{exp}', {ag_id});")
    else:
        cursor.execute(
            f"INSERT INTO storage_certificate (CertificateID, OverallStorageState, ExpirationDate, ServicingAgencyID) "
            f"VALUES ({idd}, '{state}', NULL, {ag_id});")
    mydb.commit()

def sc_update(idd, state=None, exp=None, ag_id=None):
    fields_to_update = []
    values = []

    if state:
        fields_to_update.append("OverallStorageState = %s")
        values.append(state)

    if exp:
        fields_to_update.append("ExpirationDate = %s")
        values.append(exp)

    if ag_id:
        fields_to_update.append("ServicingAgencyID = %s")
        values.append(ag_id)

    if not fields_to_update:
        return

    values.append(idd)
    query = f"UPDATE storage_certificate SET {', '.join(fields_to_update)} WHERE CertificateID = %s;"
    cursor.execute(query, values)
    mydb.commit()

def sc_delete(idd):
    cursor.execute(f"DELETE FROM storage_certificate WHERE CertificateID={idd};")
    mydb.commit()

def sc_get_all():
    cursor.execute(f"SELECT * FROM storage_certificate;")
    return cursor.fetchall()

def sc_get_more():
    cursor.execute("SELECT CertificateID, OverallStorageState, ExpirationDate, (SELECT AgencyName FROM "
    "licensed_servicing_agency WHERE storage_certificate.ServicingAgencyID=licensed_servicing_agency.AgencyID) AS Agency "
    "FROM storage_certificate;")
    return cursor.fetchall()

def nrc_add(name, loc_id, own_id, ag_id, dis_typ, stor_vol, state, was_typ, was_mass, was_vol, rad, cer_id):
    if cer_id:
        cursor.execute(
            f"INSERT INTO nws_registry_card (NWSName, LocationID, OwnerCompanyID, ServicingAgencyID, DisposalType, StorageVolume, "
            f"OverallState, WasteType, WasteMass, WasteVolume, OverallRadioactivity, CertificateID) VALUES ('{name}', {loc_id}, "
            f"{own_id}, {ag_id}, '{dis_typ}', {stor_vol}, '{state}', '{was_typ}', {was_mass}, {was_vol}, '{rad}', {cer_id});")
    else:
        cursor.execute(
            f"INSERT INTO nws_registry_card (NWSName, LocationID, OwnerCompanyID, ServicingAgencyID, DisposalType, StorageVolume, "
            f"OverallState, WasteType, WasteMass, WasteVolume, OverallRadioactivity, CertificateID) VALUES ('{name}', {loc_id}, "
            f"{own_id}, {ag_id}, '{dis_typ}', {stor_vol}, '{state}', '{was_typ}', {was_mass}, {was_vol}, '{rad}', NULL);")
    mydb.commit()

def nrc_update(name, loc_id=None, own_id=None, ag_id=None, dis_typ=None, stor_vol=None, state=None, was_typ=None, was_mass=None,
               was_vol=None, rad=None, cer_id=None):
    fields_to_update = []
    values = []

    if loc_id:
        fields_to_update.append("LocationID = %s")
        values.append(loc_id)

    if own_id:
        fields_to_update.append("OwnerCompanyID = %s")
        values.append(own_id)

    if ag_id:
        fields_to_update.append("ServicingAgencyID = %s")
        values.append(ag_id)

    if dis_typ:
        fields_to_update.append("DisposalType = %s")
        values.append(dis_typ)

    if stor_vol:
        fields_to_update.append("StorageVolume = %s")
        values.append(stor_vol)

    if state:
        fields_to_update.append("OverallState = %s")
        values.append(state)

    if was_typ:
        fields_to_update.append("WasteType = %s")
        values.append(was_typ)

    if was_mass:
        fields_to_update.append("WasteMass = %s")
        values.append(was_mass)

    if was_vol:
        fields_to_update.append("WasteVolume = %s")
        values.append(was_vol)

    if rad:
        fields_to_update.append("OverallRadioactivity = %s")
        values.append(rad)

    if cer_id:
        fields_to_update.append("CertificateID = %s")
        values.append(cer_id)

    if not fields_to_update:
        return

    values.append(name)
    query = f"UPDATE nws_registry_card SET {', '.join(fields_to_update)} WHERE NWSName = %s;"
    cursor.execute(query, values)
    mydb.commit()

def nrc_delete(name):
    cursor.execute(f"DELETE FROM nws_registry_card WHERE NWSName='{name}';")
    mydb.commit()

def nrc_get_all():
    cursor.execute("SELECT NWSName, (SELECT Region FROM nws_location WHERE nws_location.LocationID="
    "nws_registry_card.LocationID) AS Region, LocationID, OwnerCompanyID, ServicingAgencyID, DisposalType, "
    "StorageVolume, OverallState, WasteType, WasteMass, WasteVolume, OverallRadioactivity, CertificateID FROM nws_registry_card;")
    return cursor.fetchall()


def nrc_full_get():
    cursor.execute("""SELECT 
        nrc.NWSName,
        loc.Region,
        loc.District,
        CONCAT('(', loc.Latitude, ', ', loc.Longitude, ')') AS Coordinates,
        company.CompanyName AS OwnerCompanyName,
        company.CompanyOwner AS OwnerCompanyOwner,
        agency.AgencyName AS ServicingAgencyName,
        agency.AgencyOwner AS ServicingAgencyOwner,
        nrc.DisposalType,
        nrc.StorageVolume,
        nrc.OverallState,
        nrc.WasteType,
        nrc.WasteMass,
        nrc.WasteVolume,
        nrc.OverallRadioactivity,
        cert.OverallStorageState AS CertificateStorageState,
        cert.ExpirationDate AS CertificateExpirationDate
    FROM 
        nws_registry_card nrc
    LEFT JOIN 
        nws_location loc ON nrc.LocationID = loc.LocationID
    LEFT JOIN 
        nws_owner_company company ON nrc.OwnerCompanyID = company.CompanyID
    LEFT JOIN 
        licensed_servicing_agency agency ON nrc.ServicingAgencyID = agency.AgencyID
    LEFT JOIN 
        storage_certificate cert ON nrc.CertificateID = cert.CertificateID;
    """)

    return cursor.fetchall()

def noc_full_get():
    cursor.execute("""SELECT
        company.CompanyName,
        company.CompanyOwner,
        COUNT(nrc.NWSName) AS NumberOfStorages,
        COALESCE(SUM(nrc.StorageVolume), 0) AS TotalStorageVolume,
        COALESCE(SUM(nrc.WasteVolume), 0) AS TotalWasteVolume,
        COALESCE(SUM(nrc.WasteMass), 0) AS TotalWasteMass
    FROM 
        nws_owner_company company
    LEFT JOIN 
        nws_registry_card nrc ON company.CompanyID = nrc.OwnerCompanyID
    GROUP BY 
        company.CompanyID
    ORDER BY 
        company.CompanyName;
    """)

    return cursor.fetchall()

def lsa_full_get():
    cursor.execute("""SELECT
        agency.AgencyName,
        agency.AgencyOwner,
        COUNT(DISTINCT cert.CertificateID) AS NumberOfCertificates,
        COUNT(DISTINCT nrc.NWSName) AS NumberOfStorages
    FROM 
        licensed_servicing_agency agency
    LEFT JOIN 
        storage_certificate cert ON agency.AgencyID = cert.ServicingAgencyID
    LEFT JOIN 
        nws_registry_card nrc ON agency.AgencyID = nrc.ServicingAgencyID
    GROUP BY 
        agency.AgencyID
    ORDER BY 
        agency.AgencyName;
    """)

    return cursor.fetchall()

def nl_full_get():
    cursor.execute("""SELECT
        nrc.NWSName AS StorageName,
        loc.Region AS Region,
        loc.District AS District,
        loc.Latitude AS Latitude,
        loc.Longitude AS Longitude
    FROM 
        nws_registry_card nrc
    LEFT JOIN 
        nws_location loc ON nrc.LocationID = loc.LocationID
    ORDER BY 
        nrc.NWSName;
    """)

    return cursor.fetchall()

def sc_full_get():
    cursor.execute("""SELECT
        nrc.NWSName AS StorageName,
        agency.AgencyName AS CertificateGivingAgency,
        cert.OverallStorageState AS CertificateState,
        cert.ExpirationDate AS ExpirationDate
    FROM 
        nws_registry_card nrc
    LEFT JOIN 
        storage_certificate cert ON nrc.CertificateID = cert.CertificateID
    LEFT JOIN 
        licensed_servicing_agency agency ON cert.ServicingAgencyID = agency.AgencyID
    ORDER BY 
        nrc.NWSName;
    """)

    return cursor.fetchall()


class NuclearWasteStorageRegistrySystem:
    def __init__(self):
        self.root = tk.Tk()
        self.root.geometry("1200x650")
        self.root.title("Nuclear Waste Storage Registry")
        self.create_main_widgets()

    def create_main_widgets(self):
        self.bg_img = ImageTk.PhotoImage(Image.open("backgr.jpg").resize((1920, 1080)))
        self.bg_img_label = tk.Label(self.root, image=self.bg_img, bd=0)
        self.bg_img_label.place(x=0, y=0)

        self.frame_main = tk.Frame(self.root, bd=10, relief="raised")
        self.frame_main.pack(pady=(50, 0))

        self.title_label = tk.Label(self.frame_main, text="National Nuclear Waste Storage Registry", font="Times 32")
        self.title_label.pack(pady=(30, 0))

        self.frame_pics = tk.Frame(self.frame_main)
        self.frame_pics.pack(pady=(100, 50))

        self.frame_controls = tk.Frame(self.frame_main)
        self.frame_controls.pack(padx=(0, 80), pady=(0, 30))

        self.btn_view_reg = tk.Button(self.frame_controls, text="View Registry", font=('Times 15'), command=self.view_reg_choice)
        self.btn_view_reg.grid(row=0, column=0, padx=80, pady=5)

        self.btn_gen_rep = tk.Button(self.frame_controls, text="Generate Report", font=('Times 15'), command=self.view_report_preset_choice)
        self.btn_gen_rep.grid(row=0, column=1, padx=80, pady=5)

        self.btn_quit = tk.Button(self.frame_controls, text="Quit", font=('Times 15'), command=self.root.destroy)
        self.btn_quit.grid(row=0, column=2, padx=80, pady=5)

        self.view_reg_img = ImageTk.PhotoImage(Image.open("ViewRegistryImg.png").resize((200, 200)))
        self.gen_rep_img = ImageTk.PhotoImage(Image.open("GenerateReportImg.png").resize((200, 200)))
        self.quit_img = ImageTk.PhotoImage(Image.open("QuitImg.png").resize((200, 200)))

        self.view_reg_img_label = tk.Label(self.frame_pics, image=self.view_reg_img, bd=0)
        self.view_reg_img_label.grid(row=0, column=0, padx=50, pady=10)

        self.gen_rep_img_label = tk.Label(self.frame_pics, image=self.gen_rep_img, bd=0)
        self.gen_rep_img_label.grid(row=0, column=1, padx=50, pady=10)

        self.quit_img_label = tk.Label(self.frame_pics, image=self.quit_img, bd=0)
        self.quit_img_label.grid(row=0, column=2, padx=50, pady=10)

    def view_reg_choice(self):
        self.root.destroy()
        self.root = tk.Tk()
        self.root.geometry("1200x650")
        self.root.title("Available registry data")

        self.bg_img = ImageTk.PhotoImage(Image.open("backgr.jpg").resize((1920, 1080)))
        self.bg_img_label = tk.Label(self.root, image=self.bg_img, bd=0)
        self.bg_img_label.place(x=0, y=0)

        self.frame_main = tk.Frame(self.root, bd=10, relief="raised")
        self.frame_main.pack(pady=(50, 0))

        self.frame_pics1 = tk.Frame(self.frame_main)
        self.frame_pics1.grid(row=0, column=0, padx=10, pady=10)

        self.frame_choices = tk.Frame(self.frame_main)
        self.frame_choices.grid(row=0, column=1, padx=10, pady=10)

        self.frame_pics2 = tk.Frame(self.frame_main)
        self.frame_pics2.grid(row=0, column=2, padx=10, pady=10)

        self.storages_img = ImageTk.PhotoImage(Image.open("StorageImg.png").resize((100, 100)))
        self.locations_img = ImageTk.PhotoImage(Image.open("LocationImg.png").resize((100, 100)))
        self.cert_img = ImageTk.PhotoImage(Image.open("CertificateImg.png").resize((100, 100)))
        self.owner_img = ImageTk.PhotoImage(Image.open("OwnerImg.png").resize((100, 100)))
        self.servicer_img = ImageTk.PhotoImage(Image.open("ServicerImg.png").resize((100, 100)))
        self.return_img = ImageTk.PhotoImage(Image.open("ReturnImg.png").resize((100, 100)))

        self.storages_img_label = tk.Label(self.frame_pics1, image=self.storages_img, bd=0)
        self.storages_img_label.grid(row=0, column=0, padx=20, pady=30)

        self.locations_img_label = tk.Label(self.frame_pics1, image=self.locations_img, bd=0)
        self.locations_img_label.grid(row=1, column=0, padx=20, pady=30)

        self.cert_img_label = tk.Label(self.frame_pics1, image=self.cert_img, bd=0)
        self.cert_img_label.grid(row=2, column=0, padx=20, pady=30)

        self.btn_storages = tk.Button(self.frame_choices, text="Nuclear Waste Storages", font=('Times 15'), command=self.view_nrc)
        self.btn_storages.grid(row=0, column=0, padx=(0,100), pady=60)

        self.btn_owners = tk.Button(self.frame_choices, text="Owner Companies", font=('Times 15'), command=self.view_noc)
        self.btn_owners.grid(row=0, column=1, padx=(100, 0), pady=60)

        self.btn_locations = tk.Button(self.frame_choices, text="Storage Locations", font=('Times 15'), command=self.view_nl)
        self.btn_locations.grid(row=1, column=0, padx=(0, 100), pady=60)

        self.btn_servicers = tk.Button(self.frame_choices, text="Servicing Agencies", font=('Times 15'), command=self.view_lsa)
        self.btn_servicers.grid(row=1, column=1, padx=(100, 0), pady=60)

        self.btn_certs = tk.Button(self.frame_choices, text="Storage Certificates", font=('Times 15'), command=self.view_sc)
        self.btn_certs.grid(row=2, column=0, padx=(0, 100), pady=60)

        self.btn_return = tk.Button(self.frame_choices, text="Return", font=('Times 15'), command=self.return_to_main)
        self.btn_return.grid(row=2, column=1, padx=(100, 0), pady=60)

        self.owner_img_label = tk.Label(self.frame_pics2, image=self.owner_img, bd=0)
        self.owner_img_label.grid(row=0, column=0, padx=20, pady=30)

        self.servicer_img_label = tk.Label(self.frame_pics2, image=self.servicer_img, bd=0)
        self.servicer_img_label.grid(row=1, column=0, padx=20, pady=30)

        self.return_img_label = tk.Label(self.frame_pics2, image=self.return_img, bd=0)
        self.return_img_label.grid(row=2, column=0, padx=20, pady=30)

    def view_report_preset_choice(self):
        self.root.destroy()
        self.root = tk.Tk()
        self.root.geometry("1200x650")
        self.root.title("Report preset selection")

        self.bg_img = ImageTk.PhotoImage(Image.open("backgr.jpg").resize((1920, 1080)))
        self.bg_img_label = tk.Label(self.root, image=self.bg_img, bd=0)
        self.bg_img_label.place(x=0, y=0)

        self.frame_main = tk.Frame(self.root, bd=10, relief="raised")
        self.frame_main.pack(pady=(50, 0))

        self.frame_pics1 = tk.Frame(self.frame_main)
        self.frame_pics1.grid(row=0, column=0, padx=10, pady=10)

        self.frame_choices = tk.Frame(self.frame_main)
        self.frame_choices.grid(row=0, column=1, padx=10, pady=10)

        self.frame_pics2 = tk.Frame(self.frame_main)
        self.frame_pics2.grid(row=0, column=2, padx=10, pady=10)

        self.storages_img = ImageTk.PhotoImage(Image.open("StorageInfoPlusImg.png").resize((100, 100)))
        self.locations_img = ImageTk.PhotoImage(Image.open("StorageLocationDetailsImg.png").resize((100, 100)))
        self.cert_img = ImageTk.PhotoImage(Image.open("StorageCertificationImg.png").resize((100, 100)))
        self.owner_img = ImageTk.PhotoImage(Image.open("OwnerInfoPlus.png").resize((100, 100)))
        self.servicer_img = ImageTk.PhotoImage(Image.open("ServicerInfoPlus.png").resize((100, 100)))
        self.return_img = ImageTk.PhotoImage(Image.open("ReturnImg.png").resize((100, 100)))

        self.storages_img_label = tk.Label(self.frame_pics1, image=self.storages_img, bd=0)
        self.storages_img_label.grid(row=0, column=0, padx=20, pady=30)

        self.locations_img_label = tk.Label(self.frame_pics1, image=self.locations_img, bd=0)
        self.locations_img_label.grid(row=1, column=0, padx=20, pady=30)

        self.cert_img_label = tk.Label(self.frame_pics1, image=self.cert_img, bd=0)
        self.cert_img_label.grid(row=2, column=0, padx=20, pady=30)

        self.btn_pr1 = tk.Button(self.frame_choices, text="Full Storage Unit Info", font=('Times 15'), command=self.export_full_info)
        self.btn_pr1.grid(row=0, column=0, padx=(0, 80), pady=60)

        self.btn_pr2 = tk.Button(self.frame_choices, text="Owner Company Related Data", font=('Times 15'), command=self.export_noc_info)
        self.btn_pr2.grid(row=0, column=1, padx=(80, 0), pady=60)

        self.btn_pr3 = tk.Button(self.frame_choices, text="Storage Location Details", font=('Times 15'), command=self.export_nl_info)
        self.btn_pr3.grid(row=1, column=0, padx=(0, 80), pady=60)

        self.btn_pr4 = tk.Button(self.frame_choices, text="Servicing Agency Related Data", font=('Times 15'), command=self.export_lsa_info)
        self.btn_pr4.grid(row=1, column=1, padx=(80, 0), pady=60)

        self.btn_pr5 = tk.Button(self.frame_choices, text="Storage Certification Data", font=('Times 15'),command=self.export_sc_info)
        self.btn_pr5.grid(row=2, column=0, padx=(0, 80), pady=60)

        self.btn_return = tk.Button(self.frame_choices, text="Return", font=('Times 15'), command=self.return_to_main)
        self.btn_return.grid(row=2, column=1, padx=(80, 0), pady=60)

        self.owner_img_label = tk.Label(self.frame_pics2, image=self.owner_img, bd=0)
        self.owner_img_label.grid(row=0, column=0, padx=20, pady=30)

        self.servicer_img_label = tk.Label(self.frame_pics2, image=self.servicer_img, bd=0)
        self.servicer_img_label.grid(row=1, column=0, padx=20, pady=30)

        self.return_img_label = tk.Label(self.frame_pics2, image=self.return_img, bd=0)
        self.return_img_label.grid(row=2, column=0, padx=20, pady=30)

    def export_full_info(self):
        try:
            result = nrc_full_get()
            columns = [desc[0] for desc in cursor.description]

            file_path = self.export_to_excel(result, columns, "Full Unit Info Report.xlsx", "Full Unit Info")
            messagebox.showinfo("Success", f"Report exported to: {file_path}.")

        except pymysql.MySQLError as err:
            messagebox.showerror("Error", f"Couldn't add the record\nReason: {err}")

    def export_noc_info(self):
        try:
            result = noc_full_get()
            columns = [desc[0] for desc in cursor.description]

            file_path = self.export_to_excel(result, columns, "Owner Company Details Report.xlsx", "Owner Company Details")
            messagebox.showinfo("Success", f"Report exported to: {file_path}.")

        except pymysql.MySQLError as err:
            messagebox.showerror("Error", f"Couldn't add the record\nReason: {err}")

    def export_lsa_info(self):
        try:
            result = lsa_full_get()
            columns = [desc[0] for desc in cursor.description]

            file_path = self.export_to_excel(result, columns, "Servicing Agency Details Report.xlsx", "Servicing Agency Details")
            messagebox.showinfo("Success", f"Report exported to: {file_path}.")

        except pymysql.MySQLError as err:
            messagebox.showerror("Error", f"Couldn't add the record\nReason: {err}")

    def export_nl_info(self):
        try:
            result = nl_full_get()
            columns = [desc[0] for desc in cursor.description]

            file_path = self.export_to_excel(result, columns, "Storage Location Data Report.xlsx", "Storage Location Data")
            messagebox.showinfo("Success", f"Report exported to: {file_path}.")

        except pymysql.MySQLError as err:
            messagebox.showerror("Error", f"Couldn't add the record\nReason: {err}")

    def export_sc_info(self):
        try:
            result = sc_full_get()
            columns = [desc[0] for desc in cursor.description]

            file_path = self.export_to_excel(result, columns, "Certification Details Report.xlsx", "Certification Details")
            messagebox.showinfo("Success", f"Report exported to: {file_path}.")

        except pymysql.MySQLError as err:
            messagebox.showerror("Error", f"Couldn't add the record\nReason: {err}")


    def export_to_excel(self, query_result, column_names, file_name="spreadsheet.xlsx", title="Title"):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = title

        for col_num, column_name in enumerate(column_names, start=1):
            cell = sheet.cell(row=1, column=col_num)
            cell.value = column_name

        for row_num, row_data in enumerate(query_result, start=2):
            for col_num, cell_value in enumerate(row_data, start=1):
                cell = sheet.cell(row=row_num, column=col_num)
                cell.value = cell_value

        for col_num, column_name in enumerate(column_names, start=1):
            column_letter = get_column_letter(col_num)
            max_length = max(len(str(column_name)),
                             *(len(str(sheet.cell(row=row, column=col_num).value or "")) for row in
                               range(1, len(query_result) + 2)))
            sheet.column_dimensions[column_letter].width = max_length + 2

        workbook.save(file_name)
        return file_name

    def return_to_main(self):
        self.root.destroy()
        self.root = tk.Tk()
        self.root.geometry("1200x650")
        self.root.title("Nuclear Waste Storage Registry")
        self.create_main_widgets()

    def create_view_widgets(self):
        self.frame_main = tk.Frame(self.root, bd=10, relief="raised")
        self.frame_main.pack(pady=(50, 0))

        self.frame_controls = tk.Frame(self.frame_main)
        self.frame_controls.pack(pady=10)

        self.frame_filter_label = tk.Frame(self.frame_main)
        self.frame_filter_label.pack(pady=(50, 0))

        self.frame_filters = tk.Frame(self.frame_main)
        self.frame_filters.pack()

        self.filter_label = tk.Label(self.frame_filter_label, text="Data Display Filters:", font="Times 20")
        self.filter_label.grid(row=1, pady=5)

        self.frame_table = tk.Frame(self.frame_main)
        self.frame_table.pack(pady=10)

        self.btn_return = tk.Button(self.frame_controls, text="Return", font=('Times 15'), command=self.view_reg_choice)
        self.btn_return.grid(row=0, column=3, padx=5, pady=5)

    def view_nrc(self):
        self.root.destroy()
        self.root = tk.Tk()
        self.root.geometry("1200x650")
        self.root.title("NWS Registry Card")

        self.bg_img = ImageTk.PhotoImage(Image.open("backgr.jpg").resize((1920, 1080)))
        self.bg_img_label = tk.Label(self.root, image=self.bg_img, bd=0)
        self.bg_img_label.place(x=0, y=0)

        self.create_view_widgets()

        self.btn_add = tk.Button(self.frame_controls, text="Add Record", font=('Times 15'), command=self.add_nrc_record)
        self.btn_add.grid(row=0, column=0, padx=5, pady=5)

        self.btn_update = tk.Button(self.frame_controls, text="Update Record", font=('Times 15'), command=self.update_nrc_record)
        self.btn_update.grid(row=0, column=1, padx=5, pady=5)

        self.btn_remove = tk.Button(self.frame_controls, text="Remove Record", font=('Times 15'), command=self.delete_nrc_record)
        self.btn_remove.grid(row=0, column=2, padx=5, pady=5)

        self.btn_apply_filters = tk.Button(self.frame_controls, text="Apply Filters", font=('Times 15'), command=self.refresh_nrc_table)
        self.btn_apply_filters.grid(row=0, column=4, padx=(100, 5), pady=5)

        self.btn_clear_filters = tk.Button(self.frame_controls, text="Clear Filters", font=('Times 15'), command=self.clear_nrc_filters)
        self.btn_clear_filters.grid(row=0, column=5, padx=5, pady=5)

        self.filter_name = ttk.Entry(self.frame_filters, width=15)
        self.filter_name.grid(row=1, column=0, padx=3)

        self.filter_reg = ttk.Entry(self.frame_filters, width=14)
        self.filter_reg.grid(row=1, column=1, padx=3)

        self.filter_loc = ttk.Entry(self.frame_filters, width=9)
        self.filter_loc.grid(row=1, column=2, padx=3)

        self.filter_owner = ttk.Entry(self.frame_filters, width=8)
        self.filter_owner.grid(row=1, column=3, padx=3)

        self.filter_serv = ttk.Entry(self.frame_filters, width=9)
        self.filter_serv.grid(row=1, column=4, padx=3)

        self.filter_disp = ttk.Entry(self.frame_filters, width=10)
        self.filter_disp.grid(row=1, column=5, padx=3)

        self.filter_cap = ttk.Entry(self.frame_filters, width=10)
        self.filter_cap.grid(row=1, column=6, padx=3)

        self.filter_state = ttk.Entry(self.frame_filters, width=15)
        self.filter_state.grid(row=1, column=7, padx=3)

        self.filter_was_typ = ttk.Entry(self.frame_filters, width=12)
        self.filter_was_typ.grid(row=1, column=8, padx=3)

        self.filter_was_mass = ttk.Entry(self.frame_filters, width=13)
        self.filter_was_mass.grid(row=1, column=9, padx=3)

        self.filter_was_vol = ttk.Entry(self.frame_filters, width=10)
        self.filter_was_vol.grid(row=1, column=10, padx=3)

        self.filter_rad = ttk.Entry(self.frame_filters, width=12)
        self.filter_rad.grid(row=1, column=11, padx=3)

        self.filter_cer = ttk.Entry(self.frame_filters, width=10)
        self.filter_cer.grid(row=1, column=12, padx=3)

        self.tree = ttk.Treeview(self.frame_table, columns=(
            "NWSName", "Region", "LocationID", "OwnerCompanyID", "ServicingAgencyID",
            "DisposalType", "StorageVolume", "OverallState", "WasteType",
            "WasteMass", "WasteVolume", "OverallRadioactivity", "CertificateID"
        ), show="headings")

        self.tree.heading("NWSName", text="Name")
        self.tree.column("NWSName", width=98, anchor="center")
        self.tree.heading("Region", text="Region")
        self.tree.column("Region", width=92, anchor="center")
        self.tree.heading("LocationID", text="Location ID")
        self.tree.column("LocationID", width=72, anchor="center")
        self.tree.heading("OwnerCompanyID", text="Owner ID")
        self.tree.column("OwnerCompanyID", width=64, anchor="center")
        self.tree.heading("ServicingAgencyID", text="Servicer ID")
        self.tree.column("ServicingAgencyID", width=72, anchor="center")
        self.tree.heading("DisposalType", text="Disposal")
        self.tree.column("DisposalType", width=80, anchor="center")
        self.tree.heading("StorageVolume", text="Capacity")
        self.tree.column("StorageVolume", width=70, anchor="center")
        self.tree.heading("OverallState", text="State")
        self.tree.column("OverallState", width=100, anchor="center")
        self.tree.heading("WasteType", text="Waste")
        self.tree.column("WasteType", width=70, anchor="center")
        self.tree.heading("WasteMass", text="Waste kg")
        self.tree.column("WasteMass", width=100, anchor="center")
        self.tree.heading("WasteVolume", text="Waste m^3")
        self.tree.column("WasteVolume", width=70, anchor="center")
        self.tree.heading("OverallRadioactivity", text="Radiation")
        self.tree.column("OverallRadioactivity", width=80, anchor="center")
        self.tree.heading("CertificateID", text="Certificate ID")
        self.tree.column("CertificateID", width=80, anchor="center")
        self.tree.pack()

        self.refresh_nrc_table()

    def clear_nrc_filters(self):
        self.filter_name.delete(0, tk.END)
        self.filter_reg.delete(0, tk.END)
        self.filter_loc.delete(0, tk.END)
        self.filter_owner.delete(0, tk.END)
        self.filter_serv.delete(0, tk.END)
        self.filter_disp.delete(0, tk.END)
        self.filter_cap.delete(0, tk.END)
        self.filter_state.delete(0, tk.END)
        self.filter_was_typ.delete(0, tk.END)
        self.filter_was_mass.delete(0, tk.END)
        self.filter_was_vol.delete(0, tk.END)
        self.filter_rad.delete(0, tk.END)
        self.filter_cer.delete(0, tk.END)

        self.refresh_nrc_table()

    def refresh_nrc_table(self):
        for row in self.tree.get_children():
            self.tree.delete(row)

        f1 = self.filter_name.get().strip().lower()
        f2 = self.filter_reg.get().strip().lower()
        f3 = self.filter_loc.get().strip()
        f4 = self.filter_owner.get().strip()
        f5 = self.filter_serv.get().strip()
        f6 = self.filter_disp.get().strip().lower()
        f7 = self.filter_cap.get().strip()
        f8 = self.filter_state.get().strip().lower()
        f9 = self.filter_was_mass.get().strip()
        f10 = self.filter_was_typ.get().strip().lower()
        f11 = self.filter_was_vol.get().strip()
        f12 = self.filter_rad.get().strip().lower()
        f13 = self.filter_cer.get().strip()

        for record in nrc_get_all():
            if (f1 in record[0].lower() if f1 else True) and \
                    (f2 in record[1].lower() if f2 else True) and \
                    (f3 in str(record[2]) if f3 else True) and \
                    (f4 in str(record[3]) if f4 else True) and \
                    (f5 in str(record[4]) if f5 else True) and \
                    (f6 in record[5].lower() if f6 else True) and \
                    (f7 in str(record[6]) if f7 else True) and \
                    (f8 in record[7].lower() if f8 else True) and \
                    (f9 in str(record[8]) if f9 else True) and \
                    (f10 in record[9].lower() if f10 else True) and \
                    (f11 in str(record[10]) if f11 else True) and \
                    (f12 in record[11].lower() if f12 else True) and \
                    (f13 in str(record[12]) if f13 else True):
                self.tree.insert("", "end", values=record)

    def add_nrc_record(self):
        def save_record():
            try:
                name = entry_storage_name.get().strip()
                loc = int(entry_loc_id.get().strip())
                owner = int(entry_owner_id.get().strip())
                serv = int(entry_serv.get().strip())
                disp = entry_disp_typ.get().strip()
                cap = float(entry_capacity.get().strip())
                state = entry_state.get().strip()
                typ = entry_was_typ.get().strip()
                mass = float(entry_was_mass.get().strip())
                wvol = float(entry_was_vol.get().strip())
                rad = entry_rad.get().strip()
                cer = int(entry_cer.get().strip())

                if not all([name, loc, owner, serv, disp, cap, state, typ, mass, wvol, rad]):
                    messagebox.showerror("Error", "Only Certificate ID can be blank.\nAll other fields must be filled.")
                    return

                if cer:
                    cer = int(cer)

                nrc_add(name, loc, owner, serv, disp, cap, state, typ, mass, wvol, rad, cer)

                self.refresh_nrc_table()
                messagebox.showinfo("Success", "Record added successfully.")
                add_window.destroy()
            except ValueError:
                messagebox.showerror("Error", "All IDs, mass and volume must be numeric.")
            except pymysql.MySQLError as err:
                messagebox.showerror("Error", f"Couldn't add the record\nReason: {err}")

        add_window = tk.Toplevel(self.root)
        add_window.title("Add Record")

        ttk.Label(add_window, text="Storage Name:", font="Times 15").grid(row=0, column=0, padx=10, pady=10)
        entry_storage_name = ttk.Entry(add_window)
        entry_storage_name.grid(row=0, column=1, padx=10, pady=10)

        ttk.Label(add_window, text="Location ID:", font="Times 15").grid(row=1, column=0, padx=10, pady=10)
        entry_loc_id = ttk.Entry(add_window)
        entry_loc_id.grid(row=1, column=1, padx=10, pady=10)

        ttk.Label(add_window, text="Owner Company ID:", font="Times 15").grid(row=2, column=0, padx=10, pady=10)
        entry_owner_id = ttk.Entry(add_window)
        entry_owner_id.grid(row=2, column=1, padx=10, pady=10)

        ttk.Label(add_window, text="Servicing Agency ID:", font="Times 15").grid(row=3, column=0, padx=10, pady=10)
        entry_serv = ttk.Entry(add_window)
        entry_serv.grid(row=3, column=1, padx=10, pady=10)

        ttk.Label(add_window, text="Disposal Type:", font="Times 15").grid(row=4, column=0, padx=10, pady=10)
        entry_disp_typ = ttk.Entry(add_window)
        entry_disp_typ.grid(row=4, column=1, padx=10, pady=10)

        ttk.Label(add_window, text="Storage Volume (m^3):", font="Times 15").grid(row=5, column=0, padx=10, pady=10)
        entry_capacity = ttk.Entry(add_window)
        entry_capacity.grid(row=5, column=1, padx=10, pady=10)

        ttk.Label(add_window, text="Overall State:", font="Times 15").grid(row=6, column=0, padx=10, pady=10)
        entry_state = ttk.Entry(add_window)
        entry_state.grid(row=6, column=1, padx=10, pady=5)

        ttk.Label(add_window, text="Waste Type:", font="Times 15").grid(row=7, column=0, padx=10, pady=10)
        entry_was_typ = ttk.Entry(add_window)
        entry_was_typ.grid(row=7, column=1, padx=10, pady=10)

        ttk.Label(add_window, text="Waste Mass (kg):", font="Times 15").grid(row=8, column=0, padx=10, pady=10)
        entry_was_mass = ttk.Entry(add_window)
        entry_was_mass.grid(row=8, column=1, padx=10, pady=10)

        ttk.Label(add_window, text="Waste Volume (m^3):", font="Times 15").grid(row=9, column=0, padx=10, pady=10)
        entry_was_vol = ttk.Entry(add_window)
        entry_was_vol.grid(row=9, column=1, padx=10, pady=10)

        ttk.Label(add_window, text="Radioactivity:", font="Times 15").grid(row=10, column=0, padx=10, pady=10)
        entry_rad = ttk.Entry(add_window)
        entry_rad.grid(row=10, column=1, padx=10, pady=10)

        ttk.Label(add_window, text="Certificate ID:", font="Times 15").grid(row=11, column=0, padx=10, pady=10)
        entry_cer = ttk.Entry(add_window)
        entry_cer.grid(row=11, column=1, padx=10, pady=10)

        locs = ["..."]
        for record in nl_get_all():
            locs.append(str(record[0]) + " - " + record[1] + " obl, " + str(record[3]) + ", " + str(record[4]))

        sel_loc = tk.StringVar()
        sel_loc.set("...")
        loc_id_hints = tk.OptionMenu(add_window, sel_loc, *locs)
        loc_id_hints.grid(row=1, column=2)

        owners = ["..."]
        for record in noc_get_all():
            owners.append(str(record[0]) + " - " + record[1])

        sel_owner = tk.StringVar()
        sel_owner.set("...")
        owner_hints = tk.OptionMenu(add_window, sel_owner, *owners)
        owner_hints.grid(row=2, column=2)

        servs = ["..."]
        for record in lsa_get_all():
            servs.append(str(record[0]) + " - " + record[1])

        sel_serv = tk.StringVar()
        sel_serv.set("...")
        serv_hints = tk.OptionMenu(add_window, sel_serv, *servs)
        serv_hints.grid(row=3, column=2)

        cers = ["..."]
        for record in sc_get_more():
            cers.append(str(record[0]) + " - " + record[1] + ", exp: " + str(record[2]) + ", by " + record[3])

        sel_cer = tk.StringVar()
        sel_cer.set("...")
        cer_hints = tk.OptionMenu(add_window, sel_cer, *cers)
        cer_hints.grid(row=11, column=2)

        btn_save = tk.Button(add_window, text="Save Record", font="Times 15", command=save_record)
        btn_save.grid(row=0, column=3, padx=(100, 10), pady=10)

        btn_cancel = tk.Button(add_window, text="Cancel", font="Times 15", command=add_window.destroy)
        btn_cancel.grid(row=1, column=3, padx=(100, 10), pady=10)

    def update_nrc_record(self):
        def save_record():
            try:
                loc = entry_loc_id.get().strip()
                owner = entry_owner_id.get().strip()
                serv = entry_serv.get().strip()
                disp = entry_disp_typ.get().strip()
                cap = entry_capacity.get().strip()
                state = entry_state.get().strip()
                typ = entry_was_typ.get().strip()
                mass = entry_was_mass.get().strip()
                wvol = entry_was_vol.get().strip()
                rad = entry_rad.get().strip()
                cer = entry_cer.get().strip()

                if loc:
                    loc = int(loc)
                if owner:
                    owner = int(owner)
                if serv:
                    serv = int(serv)
                if cap:
                    cap = float(cap)
                if mass:
                    mass = float(mass)
                if wvol:
                    wvol = float(wvol)
                if cer:
                    cer = int(cer)

                nrc_update(upd_val, loc, owner, serv, disp, cap, state, typ, mass, wvol, rad, cer)

                self.refresh_nrc_table()
                messagebox.showinfo("Success", "Record updated successfully.")
                update_window.destroy()
            except ValueError:
                messagebox.showerror("Error", "All IDs, mass and volume must be numeric.")
            except pymysql.MySQLError as err:
                messagebox.showerror("Error", f"Couldn't add the record\nReason: {err}")

        selected_item = self.tree.focus()
        if not selected_item:
            messagebox.showwarning("Record Not Selected", "Please select a record to update.")
            return

        record = self.tree.item(selected_item, "values")
        upd_val = record[0]

        update_window = tk.Toplevel(self.root)
        update_window.title("Update Record")

        ttk.Label(update_window, text="Location ID:", font="Times 15").grid(row=1, column=0, padx=10, pady=10)
        entry_loc_id = ttk.Entry(update_window)
        entry_loc_id.grid(row=1, column=1, padx=10, pady=10)

        ttk.Label(update_window, text="Owner Company ID:", font="Times 15").grid(row=2, column=0, padx=10, pady=10)
        entry_owner_id = ttk.Entry(update_window)
        entry_owner_id.grid(row=2, column=1, padx=10, pady=10)

        ttk.Label(update_window, text="Servicing Agency ID:", font="Times 15").grid(row=3, column=0, padx=10, pady=10)
        entry_serv = ttk.Entry(update_window)
        entry_serv.grid(row=3, column=1, padx=10, pady=10)

        ttk.Label(update_window, text="Disposal Type:", font="Times 15").grid(row=4, column=0, padx=10, pady=10)
        entry_disp_typ = ttk.Entry(update_window)
        entry_disp_typ.grid(row=4, column=1, padx=10, pady=10)

        ttk.Label(update_window, text="Storage Volume (m^3):", font="Times 15").grid(row=5, column=0, padx=10, pady=10)
        entry_capacity = ttk.Entry(update_window)
        entry_capacity.grid(row=5, column=1, padx=10, pady=10)

        ttk.Label(update_window, text="Overall State:", font="Times 15").grid(row=6, column=0, padx=10, pady=10)
        entry_state = ttk.Entry(update_window)
        entry_state.grid(row=6, column=1, padx=10, pady=5)

        ttk.Label(update_window, text="Waste Type:", font="Times 15").grid(row=7, column=0, padx=10, pady=10)
        entry_was_typ = ttk.Entry(update_window)
        entry_was_typ.grid(row=7, column=1, padx=10, pady=10)

        ttk.Label(update_window, text="Waste Mass (kg):", font="Times 15").grid(row=8, column=0, padx=10, pady=10)
        entry_was_mass = ttk.Entry(update_window)
        entry_was_mass.grid(row=8, column=1, padx=10, pady=10)

        ttk.Label(update_window, text="Waste Volume (m^3):", font="Times 15").grid(row=9, column=0, padx=10, pady=10)
        entry_was_vol = ttk.Entry(update_window)
        entry_was_vol.grid(row=9, column=1, padx=10, pady=10)

        ttk.Label(update_window, text="Radioactivity:", font="Times 15").grid(row=10, column=0, padx=10, pady=10)
        entry_rad = ttk.Entry(update_window)
        entry_rad.grid(row=10, column=1, padx=10, pady=10)

        ttk.Label(update_window, text="Certificate ID:", font="Times 15").grid(row=11, column=0, padx=10, pady=10)
        entry_cer = ttk.Entry(update_window)
        entry_cer.grid(row=11, column=1, padx=10, pady=10)

        locs = ["..."]
        for record in nl_get_all():
            locs.append(str(record[0]) + " - " + record[1] + " obl, " + str(record[3]) + ", " + str(record[4]))

        sel_loc = tk.StringVar()
        sel_loc.set("...")
        loc_id_hints = tk.OptionMenu(update_window, sel_loc, *locs)
        loc_id_hints.grid(row=1, column=2)

        owners = ["..."]
        for record in noc_get_all():
            owners.append(str(record[0]) + " - " + record[1])

        sel_owner = tk.StringVar()
        sel_owner.set("...")
        owner_hints = tk.OptionMenu(update_window, sel_owner, *owners)
        owner_hints.grid(row=2, column=2)

        servs = ["..."]
        for record in lsa_get_all():
            servs.append(str(record[0]) + " - " + record[1])

        sel_serv = tk.StringVar()
        sel_serv.set("...")
        serv_hints = tk.OptionMenu(update_window, sel_serv, *servs)
        serv_hints.grid(row=3, column=2)

        cers = ["..."]
        for record in sc_get_more():
            cers.append(str(record[0]) + " - " + record[1] + ", exp: " + str(record[2]) + ", by " + record[3])

        sel_cer = tk.StringVar()
        sel_cer.set("...")
        cer_hints = tk.OptionMenu(update_window, sel_cer, *cers)
        cer_hints.grid(row=11, column=2)

        btn_save = tk.Button(update_window, text="Save Changes", font="Times 15", command=save_record)
        btn_save.grid(row=1, column=3, padx=(100, 10), pady=10)

        btn_cancel = tk.Button(update_window, text="Cancel", font="Times 15", command=update_window.destroy)
        btn_cancel.grid(row=2, column=3, padx=(100, 10), pady=10)

    def delete_nrc_record(self):
        selected_item = self.tree.focus()
        if not selected_item:
            messagebox.showwarning("Record Not Selected", "Please select a record to delete.")
            return

        record = self.tree.item(selected_item, "values")
        del_val = record[0]

        confirm = messagebox.askyesno("Confirm Deletion", "Delete selected record?")
        if confirm:
            try:
                nrc_delete(del_val)
                self.refresh_nrc_table()
                messagebox.showinfo("Success", "Record deleted successfully.")

            except pymysql.MySQLError as err:
                messagebox.showerror("Error", f"Couldn't delete the record\nReason: {err}")

    def view_nl(self):
        self.root.destroy()
        self.root = tk.Tk()
        self.root.geometry("1200x650")
        self.root.title("NWS Location")

        self.bg_img = ImageTk.PhotoImage(Image.open("backgr.jpg").resize((1920, 1080)))
        self.bg_img_label = tk.Label(self.root, image=self.bg_img, bd=0)
        self.bg_img_label.place(x=0, y=0)

        self.create_view_widgets()

        self.btn_add = tk.Button(self.frame_controls, text="Add Record", font=('Times 15'), command=self.add_nl_record)
        self.btn_add.grid(row=0, column=0, padx=5, pady=5)

        self.btn_update = tk.Button(self.frame_controls, text="Update Record", font=('Times 15'), command=self.update_nl_record)
        self.btn_update.grid(row=0, column=1, padx=5, pady=5)

        self.btn_remove = tk.Button(self.frame_controls, text="Remove Record", font=('Times 15'), command=self.delete_nl_record)
        self.btn_remove.grid(row=0, column=2, padx=5, pady=5)

        self.btn_apply_filters = tk.Button(self.frame_controls, text="Apply Filters", font=('Times 15'), command=self.refresh_nl_table)
        self.btn_apply_filters.grid(row=0, column=4, padx=(100, 5), pady=5)

        self.btn_clear_filters = tk.Button(self.frame_controls, text="Clear Filters", font=('Times 15'), command=self.clear_nl_filters)
        self.btn_clear_filters.grid(row=0, column=5, padx=5, pady=5)

        self.filter_idd = ttk.Entry(self.frame_filters, width=15)
        self.filter_idd.grid(row=1, column=0, padx=3)

        self.filter_reg = ttk.Entry(self.frame_filters, width=40)
        self.filter_reg.grid(row=1, column=1, padx=3)

        self.filter_dist = ttk.Entry(self.frame_filters, width=40)
        self.filter_dist.grid(row=1, column=2, padx=3)

        self.filter_lat = ttk.Entry(self.frame_filters, width=30)
        self.filter_lat.grid(row=1, column=3, padx=3)

        self.filter_long = ttk.Entry(self.frame_filters, width=30)
        self.filter_long.grid(row=1, column=4, padx=3)

        self.tree = ttk.Treeview(self.frame_table, columns=(
            "LocationID", "Region", "District", "Latitude", "Longitude"
        ), show="headings")

        self.tree.heading("LocationID", text="Location ID")
        self.tree.column("LocationID", width=100, anchor="center")
        self.tree.heading("Region", text="Region")
        self.tree.column("Region", width=250, anchor="center")
        self.tree.heading("District", text="District")
        self.tree.column("District", width=250, anchor="center")
        self.tree.heading("Latitude", text="Latitude")
        self.tree.column("Latitude", width=200, anchor="center")
        self.tree.heading("Longitude", text="Longitude")
        self.tree.column("Longitude", width=200, anchor="center")
        self.tree.pack()

        self.refresh_nl_table()

    def clear_nl_filters(self):
        self.filter_idd.delete(0, tk.END)
        self.filter_reg.delete(0, tk.END)
        self.filter_dist.delete(0, tk.END)
        self.filter_lat.delete(0, tk.END)
        self.filter_long.delete(0, tk.END)

        self.refresh_nl_table()

    def refresh_nl_table(self):
        for row in self.tree.get_children():
            self.tree.delete(row)

        f1 = self.filter_idd.get().strip()
        f2 = self.filter_reg.get().strip().lower()
        f3 = self.filter_dist.get().strip().lower()
        f4 = self.filter_lat.get().strip()
        f5 = self.filter_long.get().strip()

        for record in nl_get_all():
            if (f1 in str(record[0]) if f1 else True) and \
                    (f2 in record[1].lower() if f2 else True) and \
                    (f3 in record[2].lower() if f3 else True) and \
                    (f4 in str(record[3]) if f4 else True) and \
                    (f5 in str(record[4]) if f5 else True):
                self.tree.insert("", "end", values=record)

    def add_nl_record(self):
        def save_record():
            try:
                idd = int(entry_idd.get().strip())
                reg = entry_reg.get().strip()
                dist = entry_dist.get().strip()
                lat = float(entry_lat.get().strip())
                long = float(entry_long.get().strip())

                if not all([idd, reg, dist, lat, long]):
                    messagebox.showerror("Error", "All fields must be filled.")
                    return

                nl_add(idd, reg, dist, lat, long)

                self.refresh_nl_table()
                messagebox.showinfo("Success", "Record added successfully.")
                add_window.destroy()
            except ValueError:
                messagebox.showerror("Error", "ID and coordinates must be numeric.")
            except pymysql.MySQLError as err:
                messagebox.showerror("Error", f"Couldn't add the record\nReason: {err}")

        add_window = tk.Toplevel(self.root)
        add_window.title("Add Record")

        ttk.Label(add_window, text="Location ID:", font="Times 15").grid(row=0, column=0, padx=10, pady=10)
        entry_idd = ttk.Entry(add_window)
        entry_idd.grid(row=0, column=1, padx=10, pady=10)

        ttk.Label(add_window, text="Region:", font="Times 15").grid(row=1, column=0, padx=10, pady=10)
        entry_reg = ttk.Entry(add_window)
        entry_reg.grid(row=1, column=1, padx=10, pady=10)

        ttk.Label(add_window, text="District:", font="Times 15").grid(row=2, column=0, padx=10, pady=10)
        entry_dist = ttk.Entry(add_window)
        entry_dist.grid(row=2, column=1, padx=10, pady=10)

        ttk.Label(add_window, text="Latitude:", font="Times 15").grid(row=3, column=0, padx=10, pady=10)
        entry_lat = ttk.Entry(add_window)
        entry_lat.grid(row=3, column=1, padx=10, pady=10)

        ttk.Label(add_window, text="Longitude:", font="Times 15").grid(row=4, column=0, padx=10, pady=10)
        entry_long = ttk.Entry(add_window)
        entry_long.grid(row=4, column=1, padx=10, pady=10)

        btn_save = tk.Button(add_window, text="Save Record", font="Times 15", command=save_record)
        btn_save.grid(row=0, column=2, padx=(100, 10), pady=10)

        btn_cancel = tk.Button(add_window, text="Cancel", font="Times 15", command=add_window.destroy)
        btn_cancel.grid(row=1, column=2, padx=(100, 10), pady=10)

    def update_nl_record(self):
        def save_record():
            try:
                reg = entry_reg.get().strip()
                dist = entry_dist.get().strip()
                lat = entry_lat.get().strip()
                long = entry_long.get().strip()

                if lat:
                    lat = float(lat)
                if long:
                    long = float(long)

                nl_update(upd_val, reg, dist, lat, long)

                self.refresh_nl_table()
                messagebox.showinfo("Success", "Record updated successfully.")
                update_window.destroy()
            except ValueError:
                messagebox.showerror("Error", "Coordinates must be numeric.")
            except pymysql.MySQLError as err:
                messagebox.showerror("Error", f"Couldn't add the record\nReason: {err}")

        selected_item = self.tree.focus()
        if not selected_item:
            messagebox.showwarning("Record Not Selected", "Please select a record to update.")
            return

        record = self.tree.item(selected_item, "values")
        upd_val = record[0]

        update_window = tk.Toplevel(self.root)
        update_window.title("Update Record")

        ttk.Label(update_window, text="Region:", font="Times 15").grid(row=1, column=0, padx=10, pady=10)
        entry_reg = ttk.Entry(update_window)
        entry_reg.grid(row=1, column=1, padx=10, pady=10)

        ttk.Label(update_window, text="District:", font="Times 15").grid(row=2, column=0, padx=10, pady=10)
        entry_dist = ttk.Entry(update_window)
        entry_dist.grid(row=2, column=1, padx=10, pady=10)

        ttk.Label(update_window, text="Latitude:", font="Times 15").grid(row=3, column=0, padx=10, pady=10)
        entry_lat = ttk.Entry(update_window)
        entry_lat.grid(row=3, column=1, padx=10, pady=10)

        ttk.Label(update_window, text="Longitude:", font="Times 15").grid(row=4, column=0, padx=10, pady=10)
        entry_long = ttk.Entry(update_window)
        entry_long.grid(row=4, column=1, padx=10, pady=10)

        btn_save = tk.Button(update_window, text="Save Changes", font="Times 15", command=save_record)
        btn_save.grid(row=1, column=2, padx=(100, 10), pady=10)

        btn_cancel = tk.Button(update_window, text="Cancel", font="Times 15", command=update_window.destroy)
        btn_cancel.grid(row=2, column=2, padx=(100, 10), pady=10)

    def delete_nl_record(self):
        selected_item = self.tree.focus()
        if not selected_item:
            messagebox.showwarning("Record Not Selected", "Please select a record to delete.")
            return

        record = self.tree.item(selected_item, "values")
        del_val = record[0]

        confirm = messagebox.askyesno("Confirm Deletion", "Delete selected record?")
        if confirm:
            try:
                nl_delete(del_val)
                self.refresh_nl_table()
                messagebox.showinfo("Success", "Record deleted successfully.")

            except pymysql.MySQLError as err:
                messagebox.showerror("Error", f"Couldn't delete the record\nReason: {err}")

    def view_sc(self):
        self.root.destroy()
        self.root = tk.Tk()
        self.root.geometry("1200x650")
        self.root.title("Storage Certificate")

        self.bg_img = ImageTk.PhotoImage(Image.open("backgr.jpg").resize((1920, 1080)))
        self.bg_img_label = tk.Label(self.root, image=self.bg_img, bd=0)
        self.bg_img_label.place(x=0, y=0)

        self.create_view_widgets()

        self.btn_add = tk.Button(self.frame_controls, text="Add Record", font=('Times 15'), command=self.add_sc_record)
        self.btn_add.grid(row=0, column=0, padx=5, pady=5)

        self.btn_update = tk.Button(self.frame_controls, text="Update Record", font=('Times 15'), command=self.update_sc_record)
        self.btn_update.grid(row=0, column=1, padx=5, pady=5)

        self.btn_remove = tk.Button(self.frame_controls, text="Remove Record", font=('Times 15'), command=self.delete_sc_record)
        self.btn_remove.grid(row=0, column=2, padx=5, pady=5)

        self.btn_apply_filters = tk.Button(self.frame_controls, text="Apply Filters", font=('Times 15'), command=self.refresh_sc_table)
        self.btn_apply_filters.grid(row=0, column=4, padx=(100, 5), pady=5)

        self.btn_clear_filters = tk.Button(self.frame_controls, text="Clear Filters", font=('Times 15'), command=self.clear_sc_filters)
        self.btn_clear_filters.grid(row=0, column=5, padx=5, pady=5)

        self.filter_idd = ttk.Entry(self.frame_filters, width=40)
        self.filter_idd.grid(row=1, column=0, padx=3)

        self.filter_state = ttk.Entry(self.frame_filters, width=40)
        self.filter_state.grid(row=1, column=1, padx=3)

        self.filter_exp = ttk.Entry(self.frame_filters, width=40)
        self.filter_exp.grid(row=1, column=2, padx=3)

        self.filter_serv = ttk.Entry(self.frame_filters, width=40)
        self.filter_serv.grid(row=1, column=3, padx=3)

        self.tree = ttk.Treeview(self.frame_table, columns=(
            "CertificateID", "OverallStorageState", "ExpirationDate", "ServicingAgencyID"
        ), show="headings")

        self.tree.heading("CertificateID", text="Certificate ID")
        self.tree.column("CertificateID", width=250, anchor="center")
        self.tree.heading("OverallStorageState", text="Storage State")
        self.tree.column("OverallStorageState", width=250, anchor="center")
        self.tree.heading("ExpirationDate", text="Expires")
        self.tree.column("ExpirationDate", width=250, anchor="center")
        self.tree.heading("ServicingAgencyID", text="Servicer ID")
        self.tree.column("ServicingAgencyID", width=250, anchor="center")
        self.tree.pack()

        self.refresh_sc_table()

    def clear_sc_filters(self):
        self.filter_idd.delete(0, tk.END)
        self.filter_state.delete(0, tk.END)
        self.filter_exp.delete(0, tk.END)
        self.filter_serv.delete(0, tk.END)

        self.refresh_sc_table()

    def refresh_sc_table(self):
        for row in self.tree.get_children():
            self.tree.delete(row)

        f1 = self.filter_idd.get().strip()
        f2 = self.filter_state.get().strip().lower()
        f3 = self.filter_exp.get().strip()
        f4 = self.filter_serv.get().strip()

        for record in sc_get_all():
            if (f1 in str(record[0]) if f1 else True) and \
                    (f2 in record[1].lower() if f2 else True) and \
                    (f3 in str(record[2]) if f3 else True) and \
                    (f4 in str(record[3]) if f4 else True):
                self.tree.insert("", "end", values=record)

    def add_sc_record(self):
        def save_record():
            try:
                idd = int(entry_idd.get().strip())
                state = entry_state.get().strip()
                date = entry_date.get().strip()
                serv = int(entry_serv.get().strip())

                if not all([idd, state, serv]):
                    messagebox.showerror("Error", "Only expiration date can be blank.\nAll other fields must be filled.")
                    return

                sc_add(idd, state, date, serv)

                self.refresh_sc_table()
                messagebox.showinfo("Success", "Record added successfully.")
                add_window.destroy()
            except ValueError:
                messagebox.showerror("Error", "IDs must be numeric.")
            except pymysql.MySQLError as err:
                messagebox.showerror("Error", f"Couldn't add the record\nReason: {err}")

        add_window = tk.Toplevel(self.root)
        add_window.title("Add Record")

        ttk.Label(add_window, text="Certificate ID:", font="Times 15").grid(row=0, column=0, padx=10, pady=10)
        entry_idd = ttk.Entry(add_window)
        entry_idd.grid(row=0, column=1, padx=10, pady=10)

        ttk.Label(add_window, text="Storage State:", font="Times 15").grid(row=1, column=0, padx=10, pady=10)
        entry_state = ttk.Entry(add_window)
        entry_state.grid(row=1, column=1, padx=10, pady=10)

        ttk.Label(add_window, text="Expiration Date:", font="Times 15").grid(row=2, column=0, padx=10, pady=10)
        entry_date = ttk.Entry(add_window)
        entry_date.grid(row=2, column=1, padx=10, pady=10)

        ttk.Label(add_window, text="Servicer ID:", font="Times 15").grid(row=3, column=0, padx=10, pady=10)
        entry_serv = ttk.Entry(add_window)
        entry_serv.grid(row=3, column=1, padx=10, pady=10)

        servs = ["..."]
        for record in lsa_get_all():
            servs.append(str(record[0]) + " - " + record[1])

        sel_serv = tk.StringVar()
        sel_serv.set("...")
        serv_hints = tk.OptionMenu(add_window, sel_serv, *servs)
        serv_hints.grid(row=3, column=2)

        btn_save = tk.Button(add_window, text="Save Record", font="Times 15", command=save_record)
        btn_save.grid(row=0, column=3, padx=(100, 10), pady=10)

        btn_cancel = tk.Button(add_window, text="Cancel", font="Times 15", command=add_window.destroy)
        btn_cancel.grid(row=1, column=3, padx=(100, 10), pady=10)

    def update_sc_record(self):
        def save_record():
            try:
                state = entry_state.get().strip()
                date = entry_date.get().strip()
                serv = entry_serv.get().strip()

                if serv:
                    serv = int(serv)

                sc_update(upd_val, state, date, serv)

                self.refresh_sc_table()
                messagebox.showinfo("Success", "Record updated successfully.")
                update_window.destroy()
            except ValueError:
                messagebox.showerror("Error", "IDs must be numeric.")
            except pymysql.MySQLError as err:
                messagebox.showerror("Error", f"Couldn't add the record\nReason: {err}")

        selected_item = self.tree.focus()
        if not selected_item:
            messagebox.showwarning("Record Not Selected", "Please select a record to update.")
            return

        record = self.tree.item(selected_item, "values")
        upd_val = record[0]

        update_window = tk.Toplevel(self.root)
        update_window.title("Update Record")

        ttk.Label(update_window, text="Storage State:", font="Times 15").grid(row=1, column=0, padx=10, pady=10)
        entry_state = ttk.Entry(update_window)
        entry_state.grid(row=1, column=1, padx=10, pady=10)

        ttk.Label(update_window, text="Expiration Date:", font="Times 15").grid(row=2, column=0, padx=10, pady=10)
        entry_date = ttk.Entry(update_window)
        entry_date.grid(row=2, column=1, padx=10, pady=10)

        ttk.Label(update_window, text="Servicer ID:", font="Times 15").grid(row=3, column=0, padx=10, pady=10)
        entry_serv = ttk.Entry(update_window)
        entry_serv.grid(row=3, column=1, padx=10, pady=10)

        servs = ["..."]
        for record in lsa_get_all():
            servs.append(str(record[0]) + " - " + record[1])

        sel_serv = tk.StringVar()
        sel_serv.set("...")
        serv_hints = tk.OptionMenu(update_window, sel_serv, *servs)
        serv_hints.grid(row=3, column=2)

        btn_save = tk.Button(update_window, text="Save Changes", font="Times 15", command=save_record)
        btn_save.grid(row=1, column=3, padx=(100, 10), pady=10)

        btn_cancel = tk.Button(update_window, text="Cancel", font="Times 15", command=update_window.destroy)
        btn_cancel.grid(row=2, column=3, padx=(100, 10), pady=10)

    def delete_sc_record(self):
        selected_item = self.tree.focus()
        if not selected_item:
            messagebox.showwarning("Record Not Selected", "Please select a record to delete.")
            return

        record = self.tree.item(selected_item, "values")
        del_val = record[0]

        confirm = messagebox.askyesno("Confirm Deletion", "Delete selected record?")
        if confirm:
            try:
                sc_delete(del_val)
                self.refresh_sc_table()
                messagebox.showinfo("Success", "Record deleted successfully.")

            except pymysql.MySQLError as err:
                messagebox.showerror("Error", f"Couldn't delete the record\nReason: {err}")

    def view_noc(self):
        self.root.destroy()
        self.root = tk.Tk()
        self.root.geometry("1200x650")
        self.root.title("NWS Owner Company")

        self.bg_img = ImageTk.PhotoImage(Image.open("backgr.jpg").resize((1920, 1080)))
        self.bg_img_label = tk.Label(self.root, image=self.bg_img, bd=0)
        self.bg_img_label.place(x=0, y=0)

        self.create_view_widgets()

        self.btn_add = tk.Button(self.frame_controls, text="Add Record", font=('Times 15'), command=self.add_noc_record)
        self.btn_add.grid(row=0, column=0, padx=5, pady=5)

        self.btn_update = tk.Button(self.frame_controls, text="Update Record", font=('Times 15'), command=self.update_noc_record)
        self.btn_update.grid(row=0, column=1, padx=5, pady=5)

        self.btn_remove = tk.Button(self.frame_controls, text="Remove Record", font=('Times 15'), command=self.delete_noc_record)
        self.btn_remove.grid(row=0, column=2, padx=5, pady=5)

        self.btn_apply_filters = tk.Button(self.frame_controls, text="Apply Filters", font=('Times 15'), command=self.refresh_noc_table)
        self.btn_apply_filters.grid(row=0, column=4, padx=(100, 5), pady=5)

        self.btn_clear_filters = tk.Button(self.frame_controls, text="Clear Filters", font=('Times 15'), command=self.clear_noc_filters)
        self.btn_clear_filters.grid(row=0, column=5, padx=5, pady=5)

        self.filter_idd = ttk.Entry(self.frame_filters, width=23)
        self.filter_idd.grid(row=1, column=0, padx=3)

        self.filter_name = ttk.Entry(self.frame_filters, width=73)
        self.filter_name.grid(row=1, column=1, padx=3)

        self.filter_owner = ttk.Entry(self.frame_filters, width=73)
        self.filter_owner.grid(row=1, column=2, padx=3)

        self.tree = ttk.Treeview(self.frame_table, columns=(
            "CompanyID", "CompanyName", "CompanyOwner"
        ), show="headings")

        self.tree.heading("CompanyID", text="Company ID")
        self.tree.column("CompanyID", width=150, anchor="center")
        self.tree.heading("CompanyName", text="Company Name")
        self.tree.column("CompanyName", width=450, anchor="center")
        self.tree.heading("CompanyOwner", text="Company Owner")
        self.tree.column("CompanyOwner", width=450, anchor="center")
        self.tree.pack()

        self.refresh_noc_table()

    def clear_noc_filters(self):
        self.filter_idd.delete(0, tk.END)
        self.filter_name.delete(0, tk.END)
        self.filter_owner.delete(0, tk.END)

        self.refresh_noc_table()

    def refresh_noc_table(self):
        for row in self.tree.get_children():
            self.tree.delete(row)

        f1 = self.filter_idd.get().strip()
        f2 = self.filter_name.get().strip().lower()
        f3 = self.filter_owner.get().strip().lower()

        for record in noc_get_all():
            if (f1 in str(record[0]) if f1 else True) and \
                    (f2 in record[1].lower() if f2 else True) and \
                    (f3 in record[2].lower() if f3 else True):
                self.tree.insert("", "end", values=record)

    def add_noc_record(self):
        def save_record():
            try:
                idd = int(entry_idd.get().strip())
                name = entry_name.get().strip()
                owner = entry_owner.get().strip()

                if not all([idd, name, owner]):
                    messagebox.showerror("Error", "All fields must be filled.")
                    return

                noc_add(idd, name, owner)

                self.refresh_noc_table()
                messagebox.showinfo("Success", "Record added successfully.")
                add_window.destroy()
            except ValueError:
                messagebox.showerror("Error", "ID must be numeric.")
            except pymysql.MySQLError as err:
                messagebox.showerror("Error", f"Couldn't add the record\nReason: {err}")

        add_window = tk.Toplevel(self.root)
        add_window.title("Add Record")

        ttk.Label(add_window, text="Company ID:", font="Times 15").grid(row=0, column=0, padx=10, pady=10)
        entry_idd = ttk.Entry(add_window)
        entry_idd.grid(row=0, column=1, padx=10, pady=10)

        ttk.Label(add_window, text="Company Name:", font="Times 15").grid(row=1, column=0, padx=10, pady=10)
        entry_name = ttk.Entry(add_window)
        entry_name.grid(row=1, column=1, padx=10, pady=10)

        ttk.Label(add_window, text="Company Owner:", font="Times 15").grid(row=2, column=0, padx=10, pady=10)
        entry_owner = ttk.Entry(add_window)
        entry_owner.grid(row=2, column=1, padx=10, pady=10)

        btn_save = tk.Button(add_window, text="Save Record", font="Times 15", command=save_record)
        btn_save.grid(row=0, column=2, padx=(100, 10), pady=10)

        btn_cancel = tk.Button(add_window, text="Cancel", font="Times 15", command=add_window.destroy)
        btn_cancel.grid(row=1, column=2, padx=(100, 10), pady=10)

    def update_noc_record(self):
        def save_record():
            try:
                name = entry_name.get().strip()
                owner = entry_owner.get().strip()

                noc_update(upd_val, name, owner)

                self.refresh_noc_table()
                messagebox.showinfo("Success", "Record updated successfully.")
                update_window.destroy()
            except pymysql.MySQLError as err:
                messagebox.showerror("Error", f"Couldn't add the record\nReason: {err}")

        selected_item = self.tree.focus()
        if not selected_item:
            messagebox.showwarning("Record Not Selected", "Please select a record to update.")
            return

        record = self.tree.item(selected_item, "values")
        upd_val = record[0]

        update_window = tk.Toplevel(self.root)
        update_window.title("Update Record")

        ttk.Label(update_window, text="Company Name:", font="Times 15").grid(row=1, column=0, padx=10, pady=10)
        entry_name = ttk.Entry(update_window)
        entry_name.grid(row=1, column=1, padx=10, pady=10)

        ttk.Label(update_window, text="Company Owner:", font="Times 15").grid(row=2, column=0, padx=10, pady=10)
        entry_owner = ttk.Entry(update_window)
        entry_owner.grid(row=2, column=1, padx=10, pady=10)

        btn_save = tk.Button(update_window, text="Save Changes", font="Times 15", command=save_record)
        btn_save.grid(row=1, column=2, padx=(100, 10), pady=10)

        btn_cancel = tk.Button(update_window, text="Cancel", font="Times 15", command=update_window.destroy)
        btn_cancel.grid(row=2, column=2, padx=(100, 10), pady=10)

    def delete_noc_record(self):
        selected_item = self.tree.focus()
        if not selected_item:
            messagebox.showwarning("Record Not Selected", "Please select a record to delete.")
            return

        record = self.tree.item(selected_item, "values")
        del_val = record[0]

        confirm = messagebox.askyesno("Confirm Deletion", "Delete selected record?")
        if confirm:
            try:
                noc_delete(del_val)
                self.refresh_noc_table()
                messagebox.showinfo("Success", "Record deleted successfully.")

            except pymysql.MySQLError as err:
                messagebox.showerror("Error", f"Couldn't delete the record\nReason: {err}")

    def view_lsa(self):
        self.root.destroy()
        self.root = tk.Tk()
        self.root.geometry("1200x650")
        self.root.title("Licensed Servicing Agency")

        self.bg_img = ImageTk.PhotoImage(Image.open("backgr.jpg").resize((1920, 1080)))
        self.bg_img_label = tk.Label(self.root, image=self.bg_img, bd=0)
        self.bg_img_label.place(x=0, y=0)

        self.create_view_widgets()

        self.btn_add = tk.Button(self.frame_controls, text="Add Record", font=('Times 15'), command=self.add_lsa_record)
        self.btn_add.grid(row=0, column=0, padx=5, pady=5)

        self.btn_update = tk.Button(self.frame_controls, text="Update Record", font=('Times 15'), command=self.update_lsa_record)
        self.btn_update.grid(row=0, column=1, padx=5, pady=5)

        self.btn_remove = tk.Button(self.frame_controls, text="Remove Record", font=('Times 15'), command=self.delete_lsa_record)
        self.btn_remove.grid(row=0, column=2, padx=5, pady=5)

        self.btn_apply_filters = tk.Button(self.frame_controls, text="Apply Filters", font=('Times 15'), command=self.refresh_lsa_table)
        self.btn_apply_filters.grid(row=0, column=4, padx=(100, 5), pady=5)

        self.btn_clear_filters = tk.Button(self.frame_controls, text="Clear Filters", font=('Times 15'), command=self.clear_lsa_filters)
        self.btn_clear_filters.grid(row=0, column=5, padx=5, pady=5)

        self.filter_idd = ttk.Entry(self.frame_filters, width=23)
        self.filter_idd.grid(row=1, column=0, padx=3)

        self.filter_name = ttk.Entry(self.frame_filters, width=73)
        self.filter_name.grid(row=1, column=1, padx=3)

        self.filter_owner = ttk.Entry(self.frame_filters, width=73)
        self.filter_owner.grid(row=1, column=2, padx=3)

        self.tree = ttk.Treeview(self.frame_table, columns=(
            "AgencyID", "AgencyName", "AgencyOwner"
        ), show="headings")

        self.tree.heading("AgencyID", text="Agency ID")
        self.tree.column("AgencyID", width=150, anchor="center")
        self.tree.heading("AgencyName", text="Agency Name")
        self.tree.column("AgencyName", width=450, anchor="center")
        self.tree.heading("AgencyOwner", text="Agency Owner")
        self.tree.column("AgencyOwner", width=450, anchor="center")
        self.tree.pack()

        self.refresh_lsa_table()

    def clear_lsa_filters(self):
        self.filter_idd.delete(0, tk.END)
        self.filter_name.delete(0, tk.END)
        self.filter_owner.delete(0, tk.END)

        self.refresh_lsa_table()

    def refresh_lsa_table(self):
        for row in self.tree.get_children():
            self.tree.delete(row)

        f1 = self.filter_idd.get().strip()
        f2 = self.filter_name.get().strip().lower()
        f3 = self.filter_owner.get().strip().lower()

        for record in lsa_get_all():
            if (f1 in str(record[0]) if f1 else True) and \
                    (f2 in record[1].lower() if f2 else True) and \
                    (f3 in record[2].lower() if f3 else True):
                self.tree.insert("", "end", values=record)

    def add_lsa_record(self):
        def save_record():
            try:
                idd = int(entry_idd.get().strip())
                name = entry_name.get().strip()
                owner = entry_owner.get().strip()

                if not all([idd, name, owner]):
                    messagebox.showerror("Error", "All fields must be filled.")
                    return

                lsa_add(idd, name, owner)

                self.refresh_lsa_table()
                messagebox.showinfo("Success", "Record added successfully.")
                add_window.destroy()
            except ValueError:
                messagebox.showerror("Error", "ID must be numeric.")
            except pymysql.MySQLError as err:
                messagebox.showerror("Error", f"Couldn't add the record\nReason: {err}")

        add_window = tk.Toplevel(self.root)
        add_window.title("Add Record")

        ttk.Label(add_window, text="Agency ID:", font="Times 15").grid(row=0, column=0, padx=10, pady=10)
        entry_idd = ttk.Entry(add_window)
        entry_idd.grid(row=0, column=1, padx=10, pady=10)

        ttk.Label(add_window, text="Agency Name:", font="Times 15").grid(row=1, column=0, padx=10, pady=10)
        entry_name = ttk.Entry(add_window)
        entry_name.grid(row=1, column=1, padx=10, pady=10)

        ttk.Label(add_window, text="Agency Owner:", font="Times 15").grid(row=2, column=0, padx=10, pady=10)
        entry_owner = ttk.Entry(add_window)
        entry_owner.grid(row=2, column=1, padx=10, pady=10)

        btn_save = tk.Button(add_window, text="Save Record", font="Times 15", command=save_record)
        btn_save.grid(row=0, column=2, padx=(100, 10), pady=10)

        btn_cancel = tk.Button(add_window, text="Cancel", font="Times 15", command=add_window.destroy)
        btn_cancel.grid(row=1, column=2, padx=(100, 10), pady=10)

    def update_lsa_record(self):
        def save_record():
            try:
                name = entry_name.get().strip()
                owner = entry_owner.get().strip()

                lsa_update(upd_val, name, owner)

                self.refresh_lsa_table()
                messagebox.showinfo("Success", "Record updated successfully.")
                update_window.destroy()
            except pymysql.MySQLError as err:
                messagebox.showerror("Error", f"Couldn't add the record\nReason: {err}")

        selected_item = self.tree.focus()
        if not selected_item:
            messagebox.showwarning("Record Not Selected", "Please select a record to update.")
            return

        record = self.tree.item(selected_item, "values")
        upd_val = record[0]

        update_window = tk.Toplevel(self.root)
        update_window.title("Update Record")

        ttk.Label(update_window, text="Agency Name:", font="Times 15").grid(row=1, column=0, padx=10, pady=10)
        entry_name = ttk.Entry(update_window)
        entry_name.grid(row=1, column=1, padx=10, pady=10)

        ttk.Label(update_window, text="Agency Owner:", font="Times 15").grid(row=2, column=0, padx=10, pady=10)
        entry_owner = ttk.Entry(update_window)
        entry_owner.grid(row=2, column=1, padx=10, pady=10)

        btn_save = tk.Button(update_window, text="Save Changes", font="Times 15", command=save_record)
        btn_save.grid(row=1, column=2, padx=(100, 10), pady=10)

        btn_cancel = tk.Button(update_window, text="Cancel", font="Times 15", command=update_window.destroy)
        btn_cancel.grid(row=2, column=2, padx=(100, 10), pady=10)

    def delete_lsa_record(self):
        selected_item = self.tree.focus()
        if not selected_item:
            messagebox.showwarning("Record Not Selected", "Please select a record to delete.")
            return

        record = self.tree.item(selected_item, "values")
        del_val = record[0]

        confirm = messagebox.askyesno("Confirm Deletion", "Delete selected record?")
        if confirm:
            try:
                lsa_delete(del_val)
                self.refresh_lsa_table()
                messagebox.showinfo("Success", "Record deleted successfully.")

            except pymysql.MySQLError as err:
                messagebox.showerror("Error", f"Couldn't delete the record\nReason: {err}")

    def run(self):
        self.root.mainloop()


def main():
    try:
        app = NuclearWasteStorageRegistrySystem()
        app.run()

    except pymysql.MySQLError as err:
        print(f"MYSQL Error: {err}")

    finally:
        cursor.close()
        mydb.close()


if __name__ == "__main__":
    main()